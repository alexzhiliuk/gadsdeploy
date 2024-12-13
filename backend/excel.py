from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import NamedStyle, Alignment, Font

from campaign import Campaign, CampaignType


BASE_DIR = Path(__file__).resolve().parent

# На вход программа получает: список ключевых слов (3 типа), и компанию
# (группы (объявления (заголовки и описания) и конечный url, два пути и шаблон отслеживания))


class ExcelGoogleAds:

    templates = {
        CampaignType.GOOGLE: "google_template.xlsx",
        CampaignType.YANDEX: "yandex_template.xlsx",
    }

    def __init__(self, campaign: Campaign):
        self.wb = openpyxl.load_workbook(BASE_DIR / f"excel_templates/{self.templates[campaign.type]}")
        self.ws = self.wb.worksheets[0]
        self.ws2 = self.wb.worksheets[1]
        self.campaign = campaign

        self._create_styles()
        self._write()

        for col in self.ws.iter_cols():
            for cell in col:
                cell.border = self.thin_border

        self.wb.save(BASE_DIR / f"excel_templates/result.xlsx")

    def _create_styles(self):
        self.medium_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'),
                                    bottom=Side(style='medium'))
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                  bottom=Side(style='thin'))

        self.fillers = [
            PatternFill(patternType="solid", fgColor="00FFCC99"),
            PatternFill(patternType="solid", fgColor="0099CCFF"),
            PatternFill(patternType="solid", fgColor="00CCFFCC")
        ]

        self.title_font = Font(name="Arial", size=10, bold=True)
        self.count_font = Font(name="Arial", size=9, bold=True)
        self.price_font = Font(name="Arial", size=9, italic=True)
        self.default_font = Font(name="Arial", size=9)

        self.center_alignment = Alignment(horizontal='center')

    def _generate_keys(self):
        """Метод реализует запись ключей (ключевых слов) в отдельный лист файла .xlsx
        """
        row_index = 2
        for group in self.campaign.groups:
            for keyword in group.keywords:
                self.ws.cell(row_index, 1, self.campaign.name)
                self.ws.cell(row_index, 2, group.name)
                self.ws.cell(row_index, 3, str(keyword))
                row_index += 1

    def _generate_google_ads(self):
        """Метод реализует запись объявлений для google в отдельный лист файла .xlsx
        """
        row_index = 2
        for group in self.campaign.groups:
            for ad in group.ads:
                self.ws2.cell(row_index, 1, self.campaign.name)
                self.ws2.cell(row_index, 2, group.name)
                self.ws2.cell(row_index, 3, "Responsive search ad")

                col_index = 4
                for title in ad.titles[:15]:
                    self.ws2.cell(row_index, col_index, title.title)
                    self.ws2.cell(row_index, col_index + 1, str(title.position) if title.position else "")
                    col_index += 2

                col_index = 34
                for desc in ad.descriptions[:5]:
                    self.ws2.cell(row_index, col_index, desc.desc)
                    self.ws2.cell(row_index, col_index + 1, str(desc.position) if desc.position else "")
                    col_index += 2
                    if col_index > 40:
                        break

                self.ws2.cell(row_index, 42, str(ad.end_url))
                self.ws2.cell(row_index, 43, str(ad.paths[0]))
                self.ws2.cell(row_index, 44, str(ad.paths[1]))
                self.ws2.cell(row_index, 45, str(ad.tracking_url))

                row_index += 1

    def _generate_yandex_ads(self):
        """Метод реализует запись объявлений для yandex в отдельный лист файла .xlsx
        """
        row_index = 2
        for group in self.campaign.groups:
            for ad in group.ads:
                self.ws2.cell(row_index, 1, self.campaign.name)
                self.ws2.cell(row_index, 2, group.name)

                col_index = 3
                for title in ad.titles[:2]:
                    self.ws2.cell(row_index, col_index, title.title)
                    col_index += 1

                for desc in ad.descriptions[:1]:
                    self.ws2.cell(row_index, col_index, desc.desc)
                    col_index += 1

                self.ws2.cell(row_index, col_index, str(ad.end_url))

                row_index += 1

    def _write(self):
        self._generate_keys()
        if self.campaign.type == CampaignType.GOOGLE:
            self._generate_google_ads()
        elif self.campaign.type == CampaignType.YANDEX:
            self._generate_yandex_ads()

